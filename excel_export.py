import io
import pandas as pd
import openpyxl
import json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel(project_info_dict, df_all_parts, save_data_dict=None):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    COLOR_PINK = "FF99CC" 
    COLOR_GREEN = "CCFFCC" 
    COLOR_WHITE = "FFFFFF"
    
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000") 
    
    border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_outline_medium = Border(top=medium, left=medium, right=medium, bottom=medium)
    border_bottom_medium = Border(bottom=medium)
    border_right_medium = Border(right=medium)

    font_bold_lg = Font(name='Arial', size=12, bold=True)
    font_bold_std = Font(name='Arial', size=10, bold=True)
    font_std = Font(name='Arial', size=10)
    font_title_main = Font(name='Arial', size=16, bold=True, underline='single')

    df_export = df_all_parts.copy()
    chant_cols = ["Chant Avant", "Chant Arrière", "Chant Gauche", "Chant Droit"]
    for col in chant_cols:
        if col in df_export.columns:
            df_export[col] = df_export[col].map({True: 'OUI', False: 'NON', 'nan': 'NON'})
            df_export[col] = df_export[col].fillna('NON')

    # --- MODIFICATION : GROUPE PAR MATIÈRE ET ÉPAISSEUR ---
    if "Matière" in df_export.columns and "Epaisseur" in df_export.columns:
        df_export["GroupeKey"] = df_export["Matière"].astype(str) + " " + df_export["Epaisseur"].astype(str) + "mm"
    else:
        df_export["GroupeKey"] = "Défaut"

    groups = df_export["GroupeKey"].unique()

    for grp_name in groups:
        safe_name = str(grp_name).replace("/", "-").replace("?", "")[:30]
        ws = wb.create_sheet(title=safe_name)
        
        df_mat = df_export[df_export["GroupeKey"] == grp_name].reset_index(drop=True)
        
        ws.column_dimensions['A'].width = 4 
        ws.column_dimensions['B'].width = 5 
        ws.column_dimensions['C'].width = 40 
        ws.column_dimensions['D'].width = 6 
        ws.column_dimensions['E'].width = 12 
        ws.column_dimensions['F'].width = 6 
        ws.column_dimensions['G'].width = 6 
        ws.column_dimensions['H'].width = 12 
        ws.column_dimensions['I'].width = 6 
        ws.column_dimensions['J'].width = 6 
        ws.column_dimensions['K'].width = 40 

        last_col_letter = 'K'
        
        ws.merge_cells(f'F1:{last_col_letter}1')
        ws['F1'] = "FEUILLE DE DEBIT"
        ws['F1'].font = font_title_main
        ws['F1'].alignment = Alignment(horizontal='center', vertical='bottom')
        
        ws.merge_cells(f'H2:{last_col_letter}2')
        ws['H2'] = f"Date :      {project_info_dict.get('date', '')}"
        ws['H2'].font = font_bold_std
        ws['H2'].alignment = Alignment(horizontal='right', vertical='center')
        ws['H2'].border = border_bottom_medium
        
        ws.merge_cells(f'C3:{last_col_letter}3')
        ws['C3'] = project_info_dict.get('client', '').upper()
        ws['C3'].font = font_bold_lg
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C3'].border = border_thin
        
        ws['B3'] = "Client :"
        ws['B3'].font = font_bold_std
        ws['B3'].alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'C4:{last_col_letter}4')
        ws['C4'] = project_info_dict.get('ref_chantier', '')
        ws['C4'].font = font_bold_std
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C4'].border = border_thin
        
        ws['B4'] = "Réf Chantier :"
        ws['B4'].font = font_bold_std
        ws['B4'].alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'C5:{last_col_letter}5')
        ws['C5'] = project_info_dict.get('adresse_chantier', '')
        ws['C5'].font = font_std
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].border = border_thin
        
        ws['B5'] = "Adresse :"
        ws['B5'].font = font_bold_std
        ws['B5'].alignment = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[6].height = 25
        
        ws.merge_cells('B6:C6')
        ws['B6'] = "DEVIS / COMMANDE"
        ws['B6'].fill = PatternFill(start_color=COLOR_PINK, end_color=COLOR_PINK, fill_type="solid")
        ws['B6'].font = font_bold_std
        ws['B6'].alignment = Alignment(horizontal='right', vertical='center')
        ws['B6'].border = border_outline_medium
        ws['C6'].border = border_outline_medium 

        ws.merge_cells('D6:E6')
        ws['D6'] = "Date souhaitée"
        ws['D6'].fill = PatternFill(start_color=COLOR_PINK, end_color=COLOR_PINK, fill_type="solid")
        ws['D6'].font = font_bold_std
        ws['D6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D6'].border = border_outline_medium
        ws['E6'].border = border_outline_medium

        ws.merge_cells(f'F6:{last_col_letter}6')
        ws['F6'] = str(project_info_dict.get('date_souhaitee', ''))
        ws['F6'].fill = PatternFill(start_color=COLOR_PINK, end_color=COLOR_PINK, fill_type="solid")
        ws['F6'].font = font_bold_lg
        ws['F6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F6'].border = border_outline_medium
        for c_idx in range(7, 12): 
             ws.cell(row=6, column=c_idx).border = border_outline_medium

        ws.row_dimensions[7].height = 20
        
        ws.merge_cells('B7:C7')
        ws['B7'] = "Panneau / Décor :"
        ws['B7'].font = font_bold_std
        ws['B7'].alignment = Alignment(horizontal='right', vertical='center')
        ws['B7'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['B7'].border = border_outline_medium
        ws['C7'].border = border_outline_medium
        
        ws.merge_cells('D7:H7')
        # On recupère la matière propre depuis la colonne Matière (pas le groupeKey)
        mat_val = df_mat.iloc[0]['Matière'] if not df_mat.empty else ""
        ws['D7'] = mat_val
        ws['D7'].font = font_bold_std
        ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D7'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['D7'].border = border_outline_medium
        for c_idx in range(5, 9): ws.cell(row=7, column=c_idx).border = border_outline_medium

        ws['I7'] = "Epaisseur :"
        ws['I7'].font = font_bold_std
        ws['I7'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['I7'].border = border_outline_medium
        
        default_ep = 19
        if not df_mat.empty and 'Epaisseur' in df_mat.columns:
            val = df_mat.iloc[0]['Epaisseur']
            try: default_ep = float(val)
            except: default_ep = val
            
        ws.merge_cells(f'J7:{last_col_letter}7')
        ws['J7'] = default_ep
        ws['J7'].font = font_bold_std
        ws['J7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J7'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['J7'].border = border_outline_medium
        ws['K7'].border = border_outline_medium

        ws.row_dimensions[8].height = 20
        
        ws.merge_cells('B8:C8')
        ws['B8'] = "Chant :"
        ws['B8'].font = font_bold_std
        ws['B8'].alignment = Alignment(horizontal='right', vertical='center')
        ws['B8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['B8'].border = border_outline_medium
        ws['C8'].border = border_outline_medium

        ws['D8'] = "(mm)"
        ws['D8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['D8'].border = border_outline_medium

        ws['E8'] = project_info_dict.get('chant_mm', '')
        ws['E8'].font = font_bold_std
        ws['E8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['E8'].border = border_outline_medium

        ws.merge_cells('F8:H8')
        ws['F8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['F8'].border = border_outline_medium
        ws['G8'].border = border_outline_medium
        ws['H8'].border = border_outline_medium

        ws['I8'] = "Décor :"
        ws['I8'].font = font_bold_std
        ws['I8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['I8'].border = border_outline_medium

        ws.merge_cells(f'J8:{last_col_letter}8')
        ws['J8'] = project_info_dict.get('decor_chant', '')
        ws['J8'].font = font_bold_std
        ws['J8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J8'].fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        ws['J8'].border = border_outline_medium
        ws['K8'].border = border_outline_medium

        ws.merge_cells('C10:C11')
        ws['C10'] = "Référence Pièce"
        ws['C10'].font = font_bold_std
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].border = border_outline_medium
        ws['C11'].border = border_outline_medium

        ws.merge_cells('D10:D11')
        ws['D10'] = "Qté"
        ws['D10'].font = font_bold_std
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].border = border_outline_medium
        ws['D11'].border = border_outline_medium

        ws.merge_cells('K10:K11')
        ws['K10'] = "Usinage (*)"
        ws['K10'].font = font_bold_std
        ws['K10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['K10'].border = border_outline_medium
        ws['K11'].border = border_outline_medium

        ws.merge_cells('E10:E11')
        ws['E10'] = "Longueur\nen mm"
        ws['E10'].font = font_bold_std
        ws['E10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['E10'].border = border_outline_medium
        ws['E11'].border = border_outline_medium

        ws.merge_cells('H10:H11')
        ws['H10'] = "Largeur en\nmm"
        ws['H10'].font = font_bold_std
        ws['H10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['H10'].border = border_outline_medium
        ws['H11'].border = border_outline_medium

        ws.merge_cells('F10:G10')
        ws['F10'] = "Chant"
        ws['F10'].font = font_bold_std
        ws['F10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F10'].border = border_outline_medium
        ws['G10'].border = border_outline_medium
        
        ws['F11'] = "Avant"
        ws['F11'].font = font_bold_std
        ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F11'].border = border_outline_medium
        
        ws['G11'] = "Arrière"
        ws['G11'].font = font_bold_std
        ws['G11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G11'].border = border_outline_medium

        ws.merge_cells('I10:J10')
        ws['I10'] = "Chant"
        ws['I10'].font = font_bold_std
        ws['I10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['I10'].border = border_outline_medium
        ws['J10'].border = border_outline_medium
        
        ws['I11'] = "Gauche"
        ws['I11'].font = font_bold_std
        ws['I11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['I11'].border = border_outline_medium
        
        ws['J11'] = "Droit"
        ws['J11'].font = font_bold_std
        ws['J11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J11'].border = border_outline_medium

        current_row = 12
        line_number = 1
        
        for idx, row in df_mat.iterrows():
            raw_lettre = row.get("Lettre", "")
            lettre_display = raw_lettre.split('-')[-1] if '-' in str(raw_lettre) else raw_lettre

            ws.cell(row=current_row, column=1, value=line_number).font = font_bold_std 
            ws.cell(row=current_row, column=2, value=lettre_display).font = font_std
            ws.cell(row=current_row, column=3, value=row.get("Référence Pièce", "")).font = font_std
            ws.cell(row=current_row, column=4, value=row.get("Qté", 1)).font = font_std
            ws.cell(row=current_row, column=5, value=row.get("Longueur (mm)", 0)).font = font_std
            ws.cell(row=current_row, column=6, value=row.get("Chant Avant", "NON")).font = font_std
            ws.cell(row=current_row, column=7, value=row.get("Chant Arrière", "NON")).font = font_std
            ws.cell(row=current_row, column=8, value=row.get("Largeur (mm)", 0)).font = font_std
            ws.cell(row=current_row, column=9, value=row.get("Chant Gauche", "NON")).font = font_std
            ws.cell(row=current_row, column=10, value=row.get("Chant Droit", "NON")).font = font_std
            ws.cell(row=current_row, column=11, value=row.get("Usinage", "")).font = font_std

            for c_idx in range(1, 12):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.border = border_thin
                if c_idx in [3, 11]: 
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=current_row, column=1).border = Border(top=thin, bottom=thin, left=medium, right=thin)
            ws.cell(row=current_row, column=11).border = Border(top=thin, bottom=thin, left=thin, right=medium)

            current_row += 1
            line_number += 1

        while line_number <= 15:
            ws.cell(row=current_row, column=1, value=line_number).font = font_bold_std
            for c_idx in range(1, 12):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=current_row, column=1).border = Border(top=thin, bottom=thin, left=medium, right=thin)
            ws.cell(row=current_row, column=11).border = Border(top=thin, bottom=thin, left=thin, right=medium)
            
            current_row += 1
            line_number += 1

        for c_idx in range(1, 12):
            ws.cell(row=current_row-1, column=c_idx).border = Border(bottom=medium, left=thin, right=thin, top=thin)
            if c_idx == 1: ws.cell(row=current_row-1, column=c_idx).border = Border(bottom=medium, left=medium, right=thin, top=thin)
            if c_idx == 11: ws.cell(row=current_row-1, column=c_idx).border = Border(bottom=medium, left=thin, right=medium, top=thin)

    if save_data_dict:
        try:
            ws_data = wb.create_sheet(title="SaveData")
            ws_data['A1'] = json.dumps(save_data_dict, indent=2)
            ws_data.sheet_state = 'hidden'
        except: pass

    wb.save(output)
    return output.getvalue()